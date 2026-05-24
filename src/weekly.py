"""
Weekly / Monthly summary helpers.

Weekly xlsx  — merge per card (cross-day):
  Week header
  └─ Card header
     └─ Entry rows (all days merged by merchant)
     — NO day sub-headers

Monthly xlsx — merge per day:
  Month header
  └─ Card header
     └─ Day sub-header (Mon 18 May)
        └─ Entry rows (merged per merchant within that day)

Merge rule (normal entries only, installments never merged):
  Weekly:  same merchant + same card → one row, all-week amounts summed
  Monthly: same merchant + same card + same day → one row, daily amounts summed
"""

from datetime import date, timedelta
from collections import defaultdict, OrderedDict
import io

from formatter import _fmt, _merge_normals
from excel_export import (
    _card_fill, HEADER_BG, HEADER_FG,
    INSTALL_BG, ALT_BG, NORMAL_BG, BORDER,
)

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from cards import CARD_ORDER

# ── Colours ───────────────────────────────────────────────────────────────
CARD_BG = "2E5090"   # medium navy  — card section header
DAY_BG  = "D9EAD3"   # light green  — day sub-header (monthly only)
DAY_FG  = "1C4587"


# ── Date helpers ──────────────────────────────────────────────────────────

def week_bounds(ref: date) -> tuple[str, str]:
    monday = ref - timedelta(days=ref.weekday())
    sunday = monday + timedelta(days=6)
    return monday.isoformat(), sunday.isoformat()


def month_bounds(ref: date) -> tuple[str, str]:
    first = ref.replace(day=1)
    if ref.month == 12:
        last = ref.replace(day=31)
    else:
        last = ref.replace(month=ref.month + 1, day=1) - timedelta(days=1)
    return first.isoformat(), last.isoformat()


def current_week_label(ref: date) -> str:
    monday = ref - timedelta(days=ref.weekday())
    sunday = monday + timedelta(days=6)
    return f"Week {monday.strftime('%d %b')} – {sunday.strftime('%d %b %Y')}"


def current_month_label(ref: date) -> str:
    return ref.strftime("%B %Y")


def _day_label(iso_date: str) -> str:
    return date.fromisoformat(iso_date).strftime("%a %d %b")


# ── Negative amount ───────────────────────────────────────────────────────

def _neg(value: float) -> int | float:
    if value is None:
        return ""
    v = round(value, 2)
    v = int(v) if v == int(v) else v
    return -abs(v)


# ── Shared xlsx row writer ────────────────────────────────────────────────

def _write_entries(ws, row: int, entries: list[dict]) -> int:
    """Write merged entry rows into ws starting at `row`. Returns next row."""
    for i, e in enumerate(entries):
        bg = ALT_BG if i % 2 == 0 else NORMAL_BG

        if not e["is_installment"]:
            b = ws.cell(row=row, column=2, value=e["merchant_name"])
            c = ws.cell(row=row, column=3, value="")
            d = ws.cell(row=row, column=4, value=_neg(e["amount"]))
            for cell in (b, c, d):
                cell.fill      = _card_fill(bg)
                cell.font      = Font(name="Arial", size=10)
                cell.border    = BORDER
                cell.alignment = Alignment(vertical="center")
            b.alignment = Alignment(vertical="center", indent=1)
            d.alignment = Alignment(horizontal="right", vertical="center")
            d.number_format = "#,##0.##"
            row += 1
        else:
            months  = e["months"]
            full    = _neg(e["full_amount"])
            monthly = _neg(e["monthly_amount"])
            name    = e["merchant_name"]
            col = 2
            for mo in range(1, months + 1):
                b = ws.cell(row=row, column=col,     value=f"{name} {mo}/{months}")
                c = ws.cell(row=row, column=col + 1, value=full)
                d = ws.cell(row=row, column=col + 2, value=monthly)
                for cell in (b, c, d):
                    cell.fill   = _card_fill(INSTALL_BG)
                    cell.font   = Font(name="Arial", size=10)
                    cell.border = BORDER
                    cell.alignment = Alignment(vertical="center")
                b.alignment = Alignment(vertical="center", indent=1)
                c.alignment = Alignment(horizontal="right", vertical="center")
                d.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = "#,##0.##"
                d.number_format = "#,##0.##"
                col += 3
            ws.row_dimensions[row].height = 18
            row += 1
    return row


def _init_ws(wb, title: str):
    ws = wb.active
    ws.title = title[:31]
    ws.column_dimensions["A"].width = 3
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    for col_idx in range(5, 5 + 24 * 3):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12
    return ws


def _write_period_header(ws, row: int, label: str, height: int = 26) -> int:
    cell = ws.cell(row=row, column=2, value=label)
    cell.font      = Font(bold=True, color=HEADER_FG, name="Arial", size=12)
    cell.fill      = _card_fill(HEADER_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws.row_dimensions[row].height = height
    return row + 1


def _write_card_header(ws, row: int, card: str) -> int:
    ch = ws.cell(row=row, column=2, value=card)
    ch.font      = Font(bold=True, color=HEADER_FG, name="Arial", size=11)
    ch.fill      = _card_fill(CARD_BG)
    ch.alignment = Alignment(vertical="center", indent=1)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws.row_dimensions[row].height = 20
    return row + 1


def _write_day_header(ws, row: int, day_iso: str) -> int:
    dh = ws.cell(row=row, column=2, value=_day_label(day_iso))
    dh.font      = Font(bold=True, color=DAY_FG, name="Arial", size=10)
    dh.fill      = PatternFill("solid", start_color=DAY_BG, fgColor=DAY_BG)
    dh.alignment = Alignment(vertical="center", indent=2)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    ws.row_dimensions[row].height = 16
    return row + 1


# ── Weekly xlsx — merge per card (cross-day) ──────────────────────────────

def _merge_weekly(card_entries: list[dict]) -> list[dict]:
    """Merge all normal entries for a card across the entire week."""
    return _merge_normals(card_entries)


def build_weekly_xlsx(entries: list[dict], ref: date) -> bytes:
    """
    Layout: Week header → Card header → merged entry rows (no day sub-headers).
    All occurrences of the same merchant on any day are summed into one row.
    """
    wb = Workbook()
    ws = _init_ws(wb, current_week_label(ref))
    row = 1
    row = _write_period_header(ws, row, current_week_label(ref))

    # Group by card
    groups: dict[str, list[dict]] = defaultdict(list)
    for e in entries:
        groups[e["card_name"]].append(e)

    ordered = [c for c in CARD_ORDER if c in groups]
    for c in groups:
        if c not in ordered:
            ordered.append(c)

    for card in ordered:
        row = _write_card_header(ws, row, card)
        merged = _merge_weekly(groups[card])
        row = _write_entries(ws, row, merged)
        row += 1   # blank gap

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Monthly xlsx — merge per day ──────────────────────────────────────────

def build_monthly_xlsx(entries: list[dict], ref: date) -> bytes:
    """
    Layout: Month header → Card header → Day sub-header → merged entry rows.
    Entries with the same merchant on the same day are merged.
    Different days are shown separately under their own sub-header.
    """
    wb = Workbook()
    ws = _init_ws(wb, current_month_label(ref))
    row = 1
    row = _write_period_header(ws, row, current_month_label(ref))

    # Group: card → day → entries
    card_day: dict[str, dict[str, list[dict]]] = defaultdict(lambda: defaultdict(list))
    for e in entries:
        card_day[e["card_name"]][e["day"]].append(e)

    ordered = [c for c in CARD_ORDER if c in card_day]
    for c in card_day:
        if c not in ordered:
            ordered.append(c)

    for card in ordered:
        row = _write_card_header(ws, row, card)
        for day_iso in sorted(card_day[card].keys()):
            row = _write_day_header(ws, row, day_iso)
            merged = _merge_normals(card_day[card][day_iso])
            row = _write_entries(ws, row, merged)
        row += 1   # blank gap

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
