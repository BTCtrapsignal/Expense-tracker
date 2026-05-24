"""
Excel export — generates a .xlsx file from today's entries.
Each card gets its own section. Normal and installment rows follow
the exact same column layout as the TAB-separated /sum output,
so the user can paste directly into their existing spreadsheet.

Layout per entry:
  Normal:      col B = merchant name   col C = (blank)   col D = amount
  Installment: repeating group of 3 cols per month:
               [Merchant N/T]  [full_amount]  [monthly_amount]  ...
"""

import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from cards import CARD_ORDER
from formatter import _merge_normals


# ── colour palette ──────────────────────────────────────────────────────────
HEADER_BG   = "1F3864"   # dark navy — card section header
HEADER_FG   = "FFFFFF"
NORMAL_BG   = "FFFFFF"
INSTALL_BG  = "EBF3FB"   # light blue tint for installment rows
ALT_BG      = "F5F5F5"   # subtle alternating row for normal entries

THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _fmt(value: float) -> float | int:
    """Return negative number — all entries are expenses (outflows).
    Returns int if whole, else float rounded to 2dp."""
    if value is None:
        return ""
    rounded = round(value, 2)
    result  = int(rounded) if rounded == int(rounded) else rounded
    return -abs(result)


def _card_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", start_color=hex_color, fgColor=hex_color)


def build_xlsx(entries: list[dict]) -> bytes:
    """
    Build the Excel file in memory and return raw bytes.
    Caller is responsible for streaming/sending.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = date.today().strftime("%d %b %Y")

    # Fixed column widths
    ws.column_dimensions["A"].width = 3    # left margin
    ws.column_dimensions["B"].width = 26   # merchant name
    ws.column_dimensions["C"].width = 12   # blank / full amount
    ws.column_dimensions["D"].width = 12   # amount / monthly

    # Pre-set installment month columns width (up to 24 months = 8 extra groups)
    for col_idx in range(5, 5 + 24 * 3):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    # Group entries by card
    groups: dict[str, list[dict]] = {}
    for e in entries:
        groups.setdefault(e["card_name"], []).append(e)

    ordered = [c for c in CARD_ORDER if c in groups]
    for c in groups:
        if c not in ordered:
            ordered.append(c)

    row = 1  # current Excel row (1-indexed)

    for card in ordered:
        card_entries = groups[card]

        # ── Card header ─────────────────────────────────────────────────────
        cell = ws.cell(row=row, column=2, value=card)
        cell.font      = Font(bold=True, color=HEADER_FG, name="Arial", size=11)
        cell.fill      = _card_fill(HEADER_BG)
        cell.alignment = Alignment(vertical="center", indent=1)
        # Merge across a reasonable range so it looks like a full header
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
        ws.row_dimensions[row].height = 20
        row += 1

        # ── Entries — merge normal rows with same merchant ──────────────────
        card_entries = _merge_normals(card_entries)
        for i, e in enumerate(card_entries):
            bg = ALT_BG if i % 2 == 0 else NORMAL_BG

            if not e["is_installment"]:
                # Normal: B=merchant  C=(blank)  D=amount
                b = ws.cell(row=row, column=2, value=e["merchant_name"])
                c = ws.cell(row=row, column=3, value="")
                d = ws.cell(row=row, column=4, value=_fmt(e["amount"]))

                for cell in (b, c, d):
                    cell.fill      = _card_fill(bg)
                    cell.font      = Font(name="Arial", size=10)
                    cell.border    = BORDER
                    cell.alignment = Alignment(vertical="center")
                b.alignment = Alignment(vertical="center", indent=1)
                d.alignment = Alignment(horizontal="right", vertical="center")
                d.number_format = '#,##0.##'

                row += 1

            else:
                # Installment: one row, repeating 3-col groups per month
                months  = e["months"]
                full    = _fmt(e["full_amount"])
                monthly = _fmt(e["monthly_amount"])
                name    = e["merchant_name"]

                col = 2
                for mo in range(1, months + 1):
                    label = f"{name} {mo}/{months}"
                    b = ws.cell(row=row, column=col,     value=label)
                    c = ws.cell(row=row, column=col + 1, value=full)
                    d = ws.cell(row=row, column=col + 2, value=monthly)

                    for cell in (b, c, d):
                        cell.fill   = _card_fill(INSTALL_BG)
                        cell.font   = Font(name="Arial", size=10)
                        cell.border = BORDER
                        cell.alignment = Alignment(vertical="center")
                    b.alignment = Alignment(vertical="center", indent=1)
                    c.alignment = Alignment(horizontal="right", vertical="center")
                    d.alignment = Alignment(horizontal="right", vertical="center")
                    c.number_format = '#,##0.##'
                    d.number_format = '#,##0.##'

                    col += 3

                ws.row_dimensions[row].height = 18
                row += 1

        row += 1   # blank gap between card sections

    # Freeze top-left area (no freeze pane needed for plain data)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
